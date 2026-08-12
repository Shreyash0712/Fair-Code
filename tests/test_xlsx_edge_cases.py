"""Edge-case coverage for .xlsx parsing on both the Python loader
(faircode.loaders.read_table) and the browser profiler engine's
parseXLSX() (#187).

Run from the repo root:

    pytest tests/test_xlsx_edge_cases.py -q
"""

import importlib.util
import json
import subprocess
from pathlib import Path

import pytest

from faircode.loaders import read_table

_HAS_OPENPYXL = importlib.util.find_spec("openpyxl") is not None
if _HAS_OPENPYXL:
    import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent

requires_openpyxl = pytest.mark.skipif(
    not _HAS_OPENPYXL,
    reason="optional 'excel' extra not installed",
)


def _run_js_parse(workbook_path: Path):
    return subprocess.run(
        ["node", "scripts/engine-js.js", "profile-xlsx", str(workbook_path)],
        cwd=REPO_ROOT,
        capture_output=True,
        text=True,
        encoding="utf-8",
        check=False,
    )


def _save_workbook(tmp_path, workbook, filename):
    path = tmp_path / filename
    workbook.save(path)
    return path


def _headers_only_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["age", "sex", "income"])
    return wb


def _empty_workbook():
    return openpyxl.Workbook()


def _second_sheet_workbook():
    wb = openpyxl.Workbook()

    first = wb.active
    first.title = "Empty"

    second = wb.create_sheet("Data")
    second.append(["age", "sex"])
    second.append([20, "F"])

    return wb


def _hidden_first_sheet_workbook():
    wb = openpyxl.Workbook()

    first = wb.active
    first.title = "Hidden"
    first.sheet_state = "hidden"

    second = wb.create_sheet("Visible")
    second.append(["age", "sex"])
    second.append([20, "F"])

    return wb


@requires_openpyxl
@pytest.mark.parametrize(
    ("builder", "filename"),
    [
        (_headers_only_workbook, "headers_only.xlsx"),
        (_empty_workbook, "empty.xlsx"),
        (_second_sheet_workbook, "second_sheet.xlsx"),
        (_hidden_first_sheet_workbook, "hidden_first.xlsx"),
    ],
)
def test_xlsx_js_edge_cases_match_python(tmp_path, builder, filename):
    """A headers-only or empty-first-sheet workbook is a valid (if empty)
    dataset in faircode.loaders.read_table - the JS engine must agree
    instead of erroring where the Python CLI would succeed."""
    path = _save_workbook(tmp_path, builder(), filename)

    df = read_table(str(path))
    result = _run_js_parse(path)

    assert result.returncode == 0, result.stderr
    js_profile = json.loads(result.stdout)
    assert js_profile["n_rows"] == len(df)
    assert js_profile["n_cols"] == len(df.columns)


@requires_openpyxl
def test_headers_only_xlsx_python(tmp_path):
    path = _save_workbook(
        tmp_path,
        _headers_only_workbook(),
        "headers_only.xlsx",
    )

    df = read_table(str(path))

    assert list(df.columns) == ["age", "sex", "income"]
    assert df.empty


@requires_openpyxl
def test_empty_workbook_python(tmp_path):
    path = _save_workbook(
        tmp_path,
        _empty_workbook(),
        "empty.xlsx",
    )

    df = read_table(str(path))

    assert df.empty
    assert list(df.columns) == []


@requires_openpyxl
def test_empty_first_sheet_populated_second_sheet_python(tmp_path):
    path = _save_workbook(
        tmp_path,
        _second_sheet_workbook(),
        "second_sheet.xlsx",
    )

    df = read_table(str(path))

    assert df.empty


@requires_openpyxl
def test_hidden_first_sheet_python(tmp_path):
    path = _save_workbook(
        tmp_path,
        _hidden_first_sheet_workbook(),
        "hidden_first.xlsx",
    )

    df = read_table(str(path))

    assert df.empty
