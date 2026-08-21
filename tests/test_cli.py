import builtins
import importlib.util
import json
from pathlib import Path
import sys

import pandas as pd
import pytest

import faircode.cli as cli
from faircode.cli import main

REPO_ROOT = Path(__file__).resolve().parent.parent
SMALL_AUDIT = REPO_ROOT / "German Credit Lending" / "audit.yaml"

requires_openpyxl = pytest.mark.skipif(
    importlib.util.find_spec("openpyxl") is None,
    reason="optional 'excel' extra not installed",
)


def test_profile_fail_under_returns_nonzero_and_explains_score(tmp_path, capsys):
    path = tmp_path / "skewed.csv"
    path.write_text("sex\n" + "M\n" * 80 + "F\n" * 20, encoding="utf-8")

    exit_code = main(["profile", str(path), "--fail-under", "90"])

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "Representation score:" in captured.out
    assert "representation score 72/100 is below --fail-under 90" in captured.err


def test_profile_fail_under_keeps_json_output_machine_readable(tmp_path, capsys):
    path = tmp_path / "balanced.csv"
    path.write_text("sex\nM\nF\nM\nF\n", encoding="utf-8")

    exit_code = main(["profile", str(path), "--json", "--fail-under", "90"])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert json.loads(captured.out)["overall_score"] == 100
    assert captured.err == ""


def test_profile_fail_under_equal_threshold_returns_zero(tmp_path, capsys):
    path = tmp_path / "balanced.csv"
    path.write_text("sex\nM\nF\nM\nF\n", encoding="utf-8")

    exit_code = main(["profile", str(path), "--fail-under", "100"])

    captured = capsys.readouterr()

    assert exit_code == 0
    assert "Representation score: 100/100" in captured.out
    assert captured.err == ""


def test_compare_fail_on_drift_returns_nonzero_and_explains(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("sex\n" + "M\n" * 50 + "F\n" * 50, encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("sex\n" + "M\n" * 90 + "F\n" * 10, encoding="utf-8")

    exit_code = main(["compare", str(path_a), str(path_b), "--fail-on-drift"])

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "representation drift detected" in captured.err
    assert "--fail-on-drift" in captured.err


def test_compare_without_fail_on_drift_still_returns_zero(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("sex\n" + "M\n" * 50 + "F\n" * 50, encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("sex\n" + "M\n" * 90 + "F\n" * 10, encoding="utf-8")

    exit_code = main(["compare", str(path_a), str(path_b)])

    assert exit_code == 0


def test_compare_fail_on_drift_returns_zero_when_stable(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("sex\nM\nF\nM\nF\n", encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("sex\nM\nF\nM\nF\n", encoding="utf-8")

    exit_code = main(["compare", str(path_a), str(path_b), "--fail-on-drift"])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert captured.err == ""


def test_compare_applies_map_override_to_both_datasets(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("gndr\n" + "M\n" * 8 + "F\n" * 2, encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("gndr\n" + "M\n" * 5 + "F\n" * 5, encoding="utf-8")

    exit_code = main(["compare", str(path_a), str(path_b), "--json", "--map", "gndr=sex"])

    captured = capsys.readouterr()
    assert exit_code == 0
    result = json.loads(captured.out)
    assert [d["name"] for d in result["dimensions"]] == ["gndr"]
    assert [d["kind"] for d in result["dimensions"]] == ["sex"]


def test_compare_without_map_leaves_column_generically_categorical(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("gndr\n" + "M\n" * 8 + "F\n" * 2, encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("gndr\n" + "M\n" * 5 + "F\n" * 5, encoding="utf-8")

    exit_code = main(["compare", str(path_a), str(path_b), "--json"])

    captured = capsys.readouterr()
    assert exit_code == 0
    result = json.loads(captured.out)
    assert [d["kind"] for d in result["dimensions"]] == ["categorical"]


def test_map_without_equals_sign_exits_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(path), "--map", "sex_no_equals"])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert "invalid --map 'sex_no_equals', expected COL=KIND" in captured.err


def test_map_with_invalid_kind_exits_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(path), "--map", "sex=not_a_real_kind"])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert "invalid --map kind 'not_a_real_kind' for column 'sex'" in captured.err


def test_profile_missing_file_exits_2_with_clean_error(tmp_path, capsys):
    missing = tmp_path / "does_not_exist.csv"

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(missing)])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert f"error: file not found: {missing}" in captured.err


def test_profile_read_table_runtime_error_exits_2_with_clean_error(tmp_path, capsys, monkeypatch):
    path = tmp_path / "a.parquet"
    path.write_text("not a real parquet file", encoding="utf-8")

    def raise_runtime(_path):
        raise RuntimeError("reading .parquet files requires the 'pyarrow' package")

    monkeypatch.setattr(cli, "read_table", raise_runtime)

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(path)])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert "error: reading .parquet files requires the 'pyarrow' package" in captured.err


def test_profile_read_table_generic_exception_exits_2_with_clean_error(tmp_path, capsys, monkeypatch):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")

    def raise_generic(_path):
        raise ValueError("boom")

    monkeypatch.setattr(cli, "read_table", raise_generic)

    with pytest.raises(SystemExit) as exc_info:
        main(["profile", str(path)])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert f"error: could not read dataset {path}: boom" in captured.err


def test_profile_malformed_cross_returns_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")

    exit_code = main(["profile", str(path), "--cross", "onlyonecolumn"])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "--cross expects two column names: COLA,COLB" in captured.err


def test_profile_reference_missing_required_columns_returns_2_with_clean_error(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")
    ref_path = tmp_path / "ref.csv"
    ref_path.write_text("nothing,relevant\n1,2\n", encoding="utf-8")

    exit_code = main(["profile", str(path), "--reference", str(ref_path)])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "reference needs column, group, and share columns" in captured.err


def test_profile_proxy_hints_runtime_error_returns_2_with_clean_error(tmp_path, capsys, monkeypatch):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")

    def raise_runtime(_df, _dimensions):
        raise RuntimeError("proxy hints need scipy (install with: pip install faircode[proxy])")

    monkeypatch.setattr(cli, "proxy_hints", raise_runtime)

    exit_code = main(["profile", str(path), "--proxy-hints"])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "error: proxy hints need scipy" in captured.err


def test_profile_html_write_failure_returns_2_with_clean_error_not_a_traceback(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")
    bad_html_path = tmp_path / "no_such_dir" / "out.html"

    exit_code = main(["profile", str(path), "--html", str(bad_html_path)])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert f"error: could not write HTML report to {bad_html_path}" in captured.err


def test_profile_html_write_success_reports_path(tmp_path, capsys):
    path = tmp_path / "a.csv"
    path.write_text("sex\nM\nF\n", encoding="utf-8")
    html_path = tmp_path / "out.html"

    exit_code = main(["profile", str(path), "--html", str(html_path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert html_path.read_text(encoding="utf-8")
    assert f"HTML report written to {html_path}" in captured.err


def test_compare_html_write_failure_returns_2_with_clean_error_not_a_traceback(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("sex\nM\nF\n", encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("sex\nM\nF\n", encoding="utf-8")
    bad_html_path = tmp_path / "no_such_dir" / "out.html"

    exit_code = main(["compare", str(path_a), str(path_b), "--html", str(bad_html_path)])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert f"error: could not write HTML report to {bad_html_path}" in captured.err


def test_compare_html_write_success_reports_path(tmp_path, capsys):
    path_a = tmp_path / "a.csv"
    path_a.write_text("sex\nM\nF\n", encoding="utf-8")
    path_b = tmp_path / "b.csv"
    path_b.write_text("sex\nM\nF\n", encoding="utf-8")
    html_path = tmp_path / "out.html"

    exit_code = main(["compare", str(path_a), str(path_b), "--html", str(html_path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert html_path.read_text(encoding="utf-8")
    assert f"HTML report written to {html_path}" in captured.err


@requires_openpyxl
def test_profile_xlsx_reports_ignored_sheets(tmp_path, capsys):
    import openpyxl

    path = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()
    first = wb.active
    first.title = "Data"
    first.append(["sex"])
    first.append(["M"])
    first.append(["F"])
    wb.create_sheet("Notes")
    wb.create_sheet("Extra")
    wb.save(path)

    exit_code = main(["profile", str(path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert "Read sheet 'Data' - 2 other sheet(s) ignored." in captured.err


@requires_openpyxl
def test_profile_xlsx_single_sheet_stays_silent(tmp_path, capsys):
    import openpyxl

    path = tmp_path / "single_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["sex"])
    ws.append(["M"])
    ws.append(["F"])
    wb.save(path)

    exit_code = main(["profile", str(path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert "ignored" not in captured.err


# ── Benchmark subcommand tests ────────────────────────────────────────────────

def test_cli_benchmark_import_error_message(monkeypatch, capsys):
    """Lines 245-249: Catch ImportError and emit the optional install guidance."""
    real_import = builtins.__import__

    def mock_import(name, *args, **kwargs):
        if "benchmark" in name:
            raise ImportError("No module named 'sklearn'")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", mock_import)
    monkeypatch.delitem(sys.modules, "faircode.benchmark", raising=False)

    exit_code = main(["benchmark"])
    assert exit_code == 2
    captured = capsys.readouterr()
    assert "error: the benchmark command needs scikit-learn and pyyaml" in captured.err
    assert "pip install faircode[benchmark]" in captured.err


def test_cli_benchmark_paper_drift_warning_on_overrides(monkeypatch, tmp_path, capsys):
    """Lines 253-263: Stderr warning when overriding frozen default resamples/permutations."""
    pytest.importorskip("sklearn", reason="benchmark extra required")
    pytest.importorskip("fairlearn", reason="benchmark extra required")
    pytest.importorskip("yaml", reason="benchmark extra required")

    dummy_fairness = pd.DataFrame([{"audit": "German Credit Lending", "metric": "dp"}])
    dummy_perf = pd.DataFrame([{"audit": "German Credit Lending", "metric": "auc"}])

    monkeypatch.setattr(
        "faircode.benchmark.run_benchmark",
        lambda **kwargs: (dummy_fairness, dummy_perf),
    )
    monkeypatch.setattr("faircode.benchmark.write_report", lambda *args, **kwargs: None)

    out_dir = str(tmp_path / "results")
    exit_code = main([
        "benchmark",
        "--n-resamples", "50",
        "--n-permutations", "50",
        "--out", out_dir,
        "--no-plots",
    ])

    assert exit_code == 0
    captured = capsys.readouterr()
    assert "warning: --n-resamples=50, --n-permutations=50 differs from the frozen paper-run default (2000)" in captured.err
    assert f"Ran 1 audit(s), wrote 1 fairness rows and 1 performance rows to {out_dir}/" in captured.err


def test_cli_benchmark_no_manifests_found_error(tmp_path, capsys):
    """Lines 271-273: Error exit path when no audit.yaml manifests are found in --root."""
    pytest.importorskip("sklearn", reason="benchmark extra required")
    pytest.importorskip("fairlearn", reason="benchmark extra required")
    pytest.importorskip("yaml", reason="benchmark extra required")

    empty_dir = tmp_path / "empty_root"
    empty_dir.mkdir()

    exit_code = main(["benchmark", "--root", str(empty_dir)])
    assert exit_code == 2
    captured = capsys.readouterr()
    assert f"error: no audit.yaml manifests found under {empty_dir}" in captured.err


@pytest.mark.skipif(not SMALL_AUDIT.is_file(), reason="German Credit Lending fixture not found")
def test_cli_benchmark_success_run(tmp_path, capsys):
    """Lines 274-283: Full benchmark execution against the German Credit Lending fixture."""
    pytest.importorskip("sklearn", reason="benchmark extra required")
    pytest.importorskip("fairlearn", reason="benchmark extra required")
    pytest.importorskip("yaml", reason="benchmark extra required")

    out_dir = tmp_path / "results"
    exit_code = main([
        "benchmark",
        str(SMALL_AUDIT),
        "--n-resamples", "5",
        "--n-permutations", "5",
        "--out", str(out_dir),
        "--no-plots",
    ])

    assert exit_code == 0
    captured = capsys.readouterr()
    assert "Ran 1 audit(s)" in captured.err
    assert f"to {out_dir}/" in captured.err
    assert (out_dir / "results_fairness.csv").is_file()
    assert (out_dir / "results_performance.csv").is_file()
    assert (out_dir / "summary.csv").is_file()
